from __future__ import annotations

import json
from collections.abc import Iterable
from decimal import Decimal, InvalidOperation
from typing import Any

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView
from django.http import HttpRequest, HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views.decorators.http import require_http_methods
from django.utils import timezone

from openpyxl import Workbook

from .forms import ProductionEntryForm, ToolIssueForm
from .models import ProductionEntry, Tool, ToolIssue, User


class CustomLoginView(LoginView):
    template_name = "monitoring/login.html"
    redirect_authenticated_user = True

    def get_success_url(self) -> str:
        return reverse_lazy("dashboard")


@login_required
def dashboard(request: HttpRequest) -> HttpResponse:
    user: User = request.user  # type: ignore[assignment]

    if user.is_manager():
        return render(request, "monitoring/manager_dashboard.html")

    entry_form = ProductionEntryForm()
    issue_form = ToolIssueForm()
    recent_entries = (
        ProductionEntry.objects.filter(worker=user)
        .select_related("machine")
        .order_by("-recorded_at")[:20]
    )
    recent_issues = ToolIssue.objects.filter(reported_by=user).select_related("tool")[:20]

    context = {
        "entry_form": entry_form,
        "issue_form": issue_form,
        "recent_entries": recent_entries,
        "recent_issues": recent_issues,
    }
    return render(request, "monitoring/worker_dashboard.html", context)


@login_required
def create_production_entry(request: HttpRequest) -> HttpResponse:
    user: User = request.user  # type: ignore[assignment]
    if not user.is_worker():
        return HttpResponseForbidden("Доступ только для сотрудников.")

    form = ProductionEntryForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            entry = form.save(commit=False)
            entry.worker = user
            entry.save()
            messages.success(request, "Запись о выпуске деталей добавлена.")
        else:
            messages.error(request, "Не удалось сохранить данные, проверьте форму.")
    return redirect("dashboard")


@login_required
def report_tool_issue(request: HttpRequest) -> HttpResponse:
    user: User = request.user  # type: ignore[assignment]
    if not user.is_worker():
        return HttpResponseForbidden("Доступ только для сотрудников.")

    form = ToolIssueForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            issue = form.save(commit=False)
            issue.reported_by = user
            issue.save()

            tool = issue.tool
            tool.defective_stock += issue.defective_count
            tool.save(update_fields=["defective_stock", "last_updated_at"])

            messages.success(request, "Сообщение об инструменте передано руководителю.")
        else:
            messages.error(request, "Не удалось передать сообщение, проверьте форму.")
    return redirect("dashboard")


@login_required
def export_excel(request: HttpRequest) -> HttpResponse:
    user: User = request.user  # type: ignore[assignment]
    if not user.is_manager():
        return HttpResponseForbidden("Доступ только для руководителя.")

    workbook = Workbook()
    production_sheet = workbook.active
    production_sheet.title = "Производство"
    production_sheet.append(
        [
            "Дата и время",
            "Сотрудник",
            "Смена",
            "Станок",
            "Подразделение",
            "Деталь",
            "Выпуск, шт.",
            "Брак, шт.",
            "Температура, °C",
            "Вибрация, мм/с",
            "Износ, %",
            "Комментарий",
        ]
    )

    entries: Iterable[ProductionEntry] = ProductionEntry.objects.select_related(
        "worker", "machine"
    ).order_by("-recorded_at")

    for entry in entries:
        production_sheet.append(
            [
                entry.recorded_at.strftime("%Y-%m-%d %H:%M"),
                entry.worker.get_full_name() or entry.worker.username,
                entry.shift,
                entry.machine.name,
                entry.machine.subdivision,
                entry.detail_name,
                entry.parts_made,
                entry.defective_parts,
                entry.temperature_c,
                entry.vibration_mm,
                entry.tool_wear_percent,
                entry.note,
            ]
        )

    tool_sheet = workbook.create_sheet("Инструменты")
    tool_sheet.append(
        [
            "Инструмент",
            "Общий остаток",
            "Брак",
            "Мин. порог",
            "Местоположение",
            "Средний расход/день",
            "Обновлено",
        ]
    )

    for tool in Tool.objects.all():
        tool_sheet.append(
            [
                tool.name,
                tool.stock,
                tool.defective_stock,
                tool.min_threshold,
                tool.location,
                tool.avg_daily_outflow,
                tool.last_updated_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )

    issues_sheet = workbook.create_sheet("Сообщения")
    issues_sheet.append(
        [
            "Дата",
            "Сотрудник",
            "Инструмент",
            "Кол-во с браком",
            "Комментарий",
        ]
    )
    for issue in ToolIssue.objects.select_related("tool", "reported_by").all():
        issues_sheet.append(
            [
                issue.recorded_at.strftime("%Y-%m-%d %H:%M"),
                issue.reported_by.get_full_name() or issue.reported_by.username,
                issue.tool.name,
                issue.defective_count,
                issue.description,
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="quality_monitor.xlsx"'
    workbook.save(response)
    return response


def _decimal_to_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _tool_to_dict(tool: Tool) -> dict[str, Any]:
    return {
        "id": tool.pk,
        "tool_name": tool.name,
        "stock": tool.stock,
        "defective_stock": tool.defective_stock,
        "min_threshold": tool.min_threshold,
        "location": tool.location,
        "avg_daily_outflow": _decimal_to_float(tool.avg_daily_outflow),
        "updated_at": timezone.localtime(tool.last_updated_at).isoformat()
        if tool.last_updated_at
        else None,
    }


@login_required
def api_process_rows(request: HttpRequest) -> HttpResponse:
    user: User = request.user  # type: ignore[assignment]
    if not user.is_manager():
        return HttpResponseForbidden("Доступ только для руководителя.")

    rows: list[dict[str, Any]] = []
    warnings: list[str] = []

    entries = (
        ProductionEntry.objects.select_related("machine")
        .order_by("recorded_at")
        .all()
    )
    for entry in entries:
        if (
            entry.temperature_c is None
            or entry.vibration_mm is None
            or entry.tool_wear_percent is None
        ):
            warnings.append(
                f"Запись {entry.pk} от {entry.recorded_at:%Y-%m-%d %H:%M} пропущена: нет замеров."
            )
            continue
        rows.append(
            {
                "id": entry.pk,
                "t": float(entry.temperature_c),
                "v": float(entry.vibration_mm),
                "w": float(entry.tool_wear_percent),
                "defect": 1 if entry.defective_parts > 0 else 0,
                "machine": entry.machine.name,
                "machine_subdivision": entry.machine.subdivision,
                "ts": timezone.localtime(entry.recorded_at).isoformat(),
                "detail": entry.detail_name,
                "shift": entry.shift,
            }
        )

    return JsonResponse({"rows": rows, "warnings": warnings})


@login_required
def api_employee_rows(request: HttpRequest) -> HttpResponse:
    user: User = request.user  # type: ignore[assignment]
    if not user.is_manager():
        return HttpResponseForbidden("Доступ только для руководителя.")

    rows: list[dict[str, Any]] = []
    entries = (
        ProductionEntry.objects.select_related("worker", "machine")
        .order_by("recorded_at")
        .all()
    )
    for entry in entries:
        worker = entry.worker
        rows.append(
            {
                "id": worker.username.upper(),
                "name": worker.get_full_name() or worker.username,
                "shift": entry.shift or "",
                "parts_made": entry.parts_made,
                "defects": entry.defective_parts,
                "avg_temp": float(entry.temperature_c)
                if entry.temperature_c is not None
                else 0.0,
                "avg_vib": float(entry.vibration_mm)
                if entry.vibration_mm is not None
                else 0.0,
                "avg_wear": float(entry.tool_wear_percent)
                if entry.tool_wear_percent is not None
                else 0.0,
                "date": timezone.localtime(entry.recorded_at).date().isoformat(),
                "machine": entry.machine.name,
                "detail": entry.detail_name,
            }
        )
    return JsonResponse({"rows": rows})


@login_required
def api_inventory_rows(request: HttpRequest) -> HttpResponse:
    user: User = request.user  # type: ignore[assignment]
    if not user.is_manager():
        return HttpResponseForbidden("Доступ только для руководителя.")

    rows = [_tool_to_dict(tool) for tool in Tool.objects.all().order_by("name")]
    return JsonResponse({"rows": rows})


@login_required
@require_http_methods(["PATCH", "POST"])
def api_inventory_update(request: HttpRequest, pk: int) -> HttpResponse:
    user: User = request.user  # type: ignore[assignment]
    if not user.is_manager():
        return HttpResponseForbidden("Доступ только для руководителя.")

    tool = get_object_or_404(Tool, pk=pk)

    try:
        payload = json.loads(request.body or "{}")
        if not isinstance(payload, dict):
            raise ValueError("Payload must be a JSON object.")
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Некорректный JSON"}, status=400)

    errors: list[str] = []

    def parse_int(field: str, allow_none: bool = False) -> int | None:
        if field not in payload:
            return getattr(tool, field)
        value = payload.get(field)
        if value in ("", None):
            if allow_none:
                return None
            errors.append(f"Поле «{field}» не может быть пустым.")
            return getattr(tool, field)
        try:
            candidate = int(value)
            if candidate < 0:
                errors.append(f"Поле «{field}» должно быть неотрицательным.")
                return getattr(tool, field)
            return candidate
        except (TypeError, ValueError):
            errors.append(f"Поле «{field}» должно быть числом.")
            return getattr(tool, field)

    stock = parse_int("stock")
    defective_stock = parse_int("defective_stock")
    min_threshold = parse_int("min_threshold")

    avg_daily_outflow = tool.avg_daily_outflow
    if "avg_daily_outflow" in payload:
        avg_value = payload.get("avg_daily_outflow")
        if avg_value in (None, ""):
            avg_daily_outflow = None
        else:
            try:
                candidate = Decimal(str(avg_value))
                if candidate < 0:
                    errors.append("Поле «avg_daily_outflow» должно быть неотрицательным.")
                else:
                    avg_daily_outflow = candidate
            except (InvalidOperation, TypeError, ValueError):
                errors.append("Поле «avg_daily_outflow» должно быть числом.")

    location = tool.location
    if "location" in payload:
        location = str(payload.get("location") or "")

    if errors:
        return JsonResponse({"error": " ".join(errors)}, status=400)

    tool.stock = stock if stock is not None else tool.stock
    tool.defective_stock = (
        defective_stock if defective_stock is not None else tool.defective_stock
    )
    tool.min_threshold = min_threshold if min_threshold is not None else tool.min_threshold
    tool.location = location
    tool.avg_daily_outflow = avg_daily_outflow
    tool.save()

    return JsonResponse({"row": _tool_to_dict(tool)})
